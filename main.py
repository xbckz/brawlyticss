"""
Brawl Stars Match Analyzer - Enhanced with Region Filtering
Install dependencies: pip install requests openpyxl schedule pandas
"""

import requests
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from datetime import datetime
import time
import schedule
import os
import json
from collections import defaultdict

workbook = Workbook()

# Configuration
CONFIG = {
    'API_KEY': 'eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiIsImtpZCI6IjI4YTMxOGY3LTAwMDAtYTFlYi03ZmExLTJjNzQzM2M2Y2NhNSJ9.eyJpc3MiOiJzdXBlcmNlbGwiLCJhdWQiOiJzdXBlcmNlbGw6Z2FtZWFwaSIsImp0aSI6ImJlNDQzOWYwLTg3NDEtNGM0ZS05NjNmLWM1ZGYyMDQwMThkYSIsImlhdCI6MTc2MzU4NzU3MCwic3ViIjoiZGV2ZWxvcGVyLzNiYzViZTBkLTUzZDItYzc4MC04OWEwLTIyY2ZiYTY3MTU5NyIsInNjb3BlcyI6WyJicmF3bHN0YXJzIl0sImxpbWl0cyI6W3sidGllciI6ImRldmVsb3Blci9zaWx2ZXIiLCJ0eXBlIjoidGhyb3R0bGluZyJ9LHsiY2lkcnMiOlsiODIuOTkuMTgwLjExNCJdLCJ0eXBlIjoiY2xpZW50In1dfQ.ewNzQF1_BQHWwSdQluOMV8n22ZLZ9i9_1kgw5sAUwoE00LGElThgQgHOjEZosnXjfS0tcb_IRin7q5bx6tDDDA',
    'CHECK_INTERVAL_MINUTES': 1,
    'TEAMS_FILE': 'teams.xlsx',
    'MATCHES_FILE': 'matches.xlsx',
    'STATS_FILE': 'statistics.xlsx',
    'MIN_PLAYERS_PER_TEAM': 2,
    'REGIONS': ['NA', 'EU', 'LATAM', 'APAC', 'SA', 'OCE'],
    'MODES': ['Brawl Ball', 'Heist', 'Bounty', 'Knockout', 'Gem Grab', 'Hot Zone']
}

# Global storage
processed_match_ids = set()
matches_data = []
player_names_map = {}

def create_teams_template():
    """Create a formatted teams.xlsx template with dropdown menus"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"
    
    # Headers with styling
    headers = [
        'Team Name', 'Region', 
        'Player 1 ID', 'Player 1 Name',
        'Player 2 ID', 'Player 2 Name',
        'Player 3 ID', 'Player 3 Name',
        'Notes'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data validation for Region column (dropdown menu)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(CONFIG["REGIONS"])}"',
        allow_blank=False
    )
    dv.error = 'Please select a valid region'
    dv.errorTitle = 'Invalid Region'
    dv.prompt = 'Select a region from the dropdown'
    dv.promptTitle = 'Region Selection'
    
    ws.add_data_validation(dv)
    dv.add(f'B2:B1000')
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 25
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions")
    instructions = [
        ["Brawl Stars Match Analyzer - Instructions"],
        [""],
        ["1. Fill in the Teams sheet with your team information"],
        ["2. Team Name: Give your team a unique name"],
        ["3. Region: Select from the dropdown menu (NA, EU, LATAM, APAC, SA, OCE)"],
        ["4. Player IDs: Enter player IDs without the # symbol (e.g., 898L2Q22)"],
        ["5. Player Names: Enter the player's in-game name"],
        ["6. You need at least 2 players per team for matches to be detected"],
        ["7. Save the file and run the Python script"],
        [""],
        ["The script will:"],
        ["- Check for matches every 30 minutes"],
        ["- Create matches.xlsx with all detected team vs team matches"],
        ["- Create statistics.xlsx with win rates, brawler picks, and more"],
        [""],
        ["Notes:"],
        ["- Player IDs can be found in-game or via the Brawl Stars API"],
        ["- Make sure your API key is set in the Python script"],
        ["- The script must be running for automatic updates"],
    ]
    
    for row_idx, row in enumerate(instructions, 1):
        ws_instructions.cell(row=row_idx, column=1, value=row[0])
        if row_idx == 1:
            ws_instructions.cell(row=row_idx, column=1).font = Font(size=14, bold=True)
        elif row_idx in [11, 16]:
            ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True)
    
    ws_instructions.column_dimensions['A'].width = 80
    
    wb.save(CONFIG['TEAMS_FILE'])
    print(f"Created template: {CONFIG['TEAMS_FILE']}")
    print(f"Please fill in your teams and save the file")


def load_teams():
    """Load teams from Excel file and build player names map"""
    global player_names_map
    
    if not os.path.exists(CONFIG['TEAMS_FILE']):
        print(f"  {CONFIG['TEAMS_FILE']} not found. Creating template...")
        create_teams_template()
        return []
    
    try:
        df = pd.read_excel(CONFIG['TEAMS_FILE'], sheet_name='Teams')
        teams = []
        player_names_map = {}
        
        for _, row in df.iterrows():
            players = []
            for i in range(1, 4):
                player_id = str(row.get(f'Player {i} ID', '')).strip()
                player_name = str(row.get(f'Player {i} Name', '')).strip()
                
                if player_id and player_id.lower() not in ['nan', 'none', '']:
                    player_tag = player_id if player_id.startswith('#') else f'#{player_id}'
                    display_name = player_name if player_name and player_name.lower() not in ['nan', 'none', ''] else player_id
                    
                    players.append({
                        'id': player_tag,
                        'name': display_name
                    })
                    
                    player_names_map[player_tag] = display_name
            
            if len(players) >= CONFIG['MIN_PLAYERS_PER_TEAM']:
                region = str(row['Region']).strip().upper()
                if region in ['NAN', '']:
                    region = 'NA'
                
                teams.append({
                    'name': str(row['Team Name']),
                    'region': region,
                    'players': players
                })
        
        print(f"Loaded {len(teams)} teams from {CONFIG['TEAMS_FILE']}")
        print(f"Player names mapped: {len(player_names_map)} players")
        return teams
    except Exception as e:
        print(f"Error loading teams: {e}")
        return []


def load_existing_matches():
    """Load existing matches to avoid duplicates"""
    global matches_data, processed_match_ids
    
    if os.path.exists(CONFIG['MATCHES_FILE']):
        try:
            df = pd.read_excel(CONFIG['MATCHES_FILE'])
            matches_data = df.to_dict('records')
            processed_match_ids = set(df['match_id'].tolist())
            print(f"Loaded {len(matches_data)} existing matches")
        except Exception as e:
            print(f"Could not load existing matches: {e}")


def fetch_player_battles(player_id):
    """Fetch battle log for a player"""
    clean_id = player_id.replace('#', '')
    url = f"https://api.brawlstars.com/v1/players/%23{clean_id}/battlelog"
    
    headers = {
        'Authorization': f'Bearer {CONFIG["API_KEY"]}'
    }
    
    try:
        response = requests.get(url, headers=headers, timeout=10)
        if response.status_code == 200:
            return response.json().get('items', [])
        else:
            print(f"Failed to fetch {player_id}: Status {response.status_code}")
            return []
    except Exception as e:
        print(f"Error fetching {player_id}: {e}")
        return []


def analyze_match(match, teams):
    """Check if match is a team vs team match"""
    if not match.get('battle') or not match['battle'].get('teams'):
        return None
    
    battle_teams = match['battle']['teams']
    if len(battle_teams) != 2:
        return None
    
    team1_players = [p['tag'] for p in battle_teams[0]]
    team2_players = [p['tag'] for p in battle_teams[1]]
    
    detected_teams = []
    
    for team in teams:
        team_ids = [p['id'] for p in team['players']]
        
        team1_count = sum(1 for pid in team1_players if pid in team_ids)
        team2_count = sum(1 for pid in team2_players if pid in team_ids)
        
        if team1_count >= CONFIG['MIN_PLAYERS_PER_TEAM']:
            detected_teams.append({'team': team, 'side': 0, 'count': team1_count})
        elif team2_count >= CONFIG['MIN_PLAYERS_PER_TEAM']:
            detected_teams.append({'team': team, 'side': 1, 'count': team2_count})
    
    if len(detected_teams) >= 2:
        team1_det = next((t for t in detected_teams if t['side'] == 0), None)
        team2_det = next((t for t in detected_teams if t['side'] == 1), None)
        
        if team1_det and team2_det:
            return {
                'team1': team1_det['team'],
                'team2': team2_det['team'],
                'team1_players': battle_teams[0],
                'team2_players': battle_teams[1],
                'result': match['battle']['result'],
                'star_player': match['battle'].get('starPlayer'),
                'battle_time': match['battleTime'],
                'mode': match['event']['mode'],
                'map': match['event']['map'],
                'type': match['battle']['type'],
                'duration': match['battle']['duration'],
                'event_id': match['event']['id']
            }
    
    return None


def analyze_matches():
    """Main analysis function"""
    print('\nStarting match analysis...')
    print('=' * 60)
    
    teams = load_teams()
    if len(teams) < 2:
        print('Need at least 2 teams in teams.xlsx')
        return
    
    print(f"Analyzing {len(teams)} teams...")
    
    new_matches_count = 0
    all_battles = {}
    
    for team in teams:
        for player in team['players']:
            print(f"  Fetching battles for {player['name']} ({player['id']})...")
            battles = fetch_player_battles(player['id'])
            if battles:
                all_battles[player['id']] = battles
            time.sleep(0.1)
    
    for player_id, battles in all_battles.items():
        for match in battles:
            match_id = f"{match['battleTime']}-{match['event']['id']}"
            
            if match_id not in processed_match_ids:
                analysis = analyze_match(match, teams)
                
                if analysis:
                    processed_match_ids.add(match_id)
                    new_matches_count += 1
                    
                    match_record = {
                        'match_id': match_id,
                        'timestamp': datetime.fromisoformat(analysis['battle_time'].replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S'),
                        'team1_name': analysis['team1']['name'],
                        'team1_region': analysis['team1']['region'],
                        'team2_name': analysis['team2']['name'],
                        'team2_region': analysis['team2']['region'],
                        'result': analysis['result'],
                        'winner': analysis['team1']['name'] if analysis['result'] == 'victory' else analysis['team2']['name'],
                        'mode': analysis['mode'],
                        'map': analysis['map'],
                        'match_type': analysis['type'],
                        'duration_seconds': analysis['duration'],
                    }
                    
                    for i in range(3):
                        match_record[f'team1_player{i+1}'] = analysis['team1_players'][i]['name']
                        match_record[f'team1_player{i+1}_brawler'] = analysis['team1_players'][i]['brawler']['name']
                        match_record[f'team1_player{i+1}_tag'] = analysis['team1_players'][i]['tag']
                        match_record[f'team2_player{i+1}'] = analysis['team2_players'][i]['name']
                        match_record[f'team2_player{i+1}_brawler'] = analysis['team2_players'][i]['brawler']['name']
                        match_record[f'team2_player{i+1}_tag'] = analysis['team2_players'][i]['tag']
                    
                    if analysis['star_player']:
                        match_record['star_player'] = f"{analysis['star_player']['name']} ({analysis['star_player']['brawler']['name']})"
                        match_record['star_player_tag'] = analysis['star_player']['tag']
                    else:
                        match_record['star_player'] = 'N/A'
                        match_record['star_player_tag'] = 'N/A'
                    
                    match_record['analyzed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    
                    matches_data.append(match_record)
                    print(f"New match: {analysis['team1']['name']} vs {analysis['team2']['name']} - {analysis['result']}")
    
    print(f"\nFound {new_matches_count} new match(es)")
    
    if new_matches_count > 0:
        save_to_excel()


def save_to_excel():
    """Save matches and statistics to Excel files"""
    try:
        if matches_data:
            df_matches = pd.DataFrame(matches_data)
            df_matches.to_excel(CONFIG['MATCHES_FILE'], index=False, sheet_name='Matches')
            print(f"Saved {len(matches_data)} matches to {CONFIG['MATCHES_FILE']}")
        
        save_statistics()
        
    except Exception as e:
        print(f"Error saving to Excel: {e}")


def calculate_advanced_statistics():
    """Calculate comprehensive statistics from matches"""
    if not matches_data:
        return None
    
    df = pd.DataFrame(matches_data)
    
    stats = {
        'team_stats': defaultdict(lambda: {
            'region': '',
            'matches': 0,
            'wins': 0,
            'losses': 0,
            'brawlers': defaultdict(lambda: {'picks': 0, 'wins': 0}),
            'players': defaultdict(lambda: {
                'brawlers': defaultdict(lambda: {'picks': 0, 'wins': 0}),
                'star_player': 0,
                'matches': 0
            }),
            'maps': defaultdict(lambda: defaultdict(lambda: {'wins': 0, 'losses': 0}))
        }),
        'brawler_stats': defaultdict(lambda: {'picks': 0, 'wins': 0}),
        'map_stats': defaultdict(lambda: defaultdict(lambda: {
            'matches': 0,
            'teams': defaultdict(lambda: {
                'wins': 0, 
                'losses': 0,
                'brawlers': defaultdict(lambda: {'picks': 0, 'wins': 0})
            })
        }))
    }
    
    for _, match in df.iterrows():
        winner = match['winner']
        loser = match['team1_name'] if winner == match['team2_name'] else match['team2_name']
        winner_region = match['team1_region'] if winner == match['team1_name'] else match['team2_region']
        loser_region = match['team2_region'] if winner == match['team1_name'] else match['team1_region']
        
        mode = match['mode']
        map_name = match['map']
        
        for team, is_winner, region in [(winner, True, winner_region), (loser, False, loser_region)]:
            stats['team_stats'][team]['region'] = region
            stats['team_stats'][team]['matches'] += 1
            if is_winner:
                stats['team_stats'][team]['wins'] += 1
                stats['team_stats'][team]['maps'][mode][map_name]['wins'] += 1
            else:
                stats['team_stats'][team]['losses'] += 1
                stats['team_stats'][team]['maps'][mode][map_name]['losses'] += 1
            
            team_prefix = 'team1' if (is_winner and match['result'] == 'victory') or (not is_winner and match['result'] != 'victory') else 'team2'
            
            for i in range(1, 4):
                player_name = match[f'{team_prefix}_player{i}']
                player_tag = match[f'{team_prefix}_player{i}_tag']
                brawler = match[f'{team_prefix}_player{i}_brawler']
                
                stats['team_stats'][team]['brawlers'][brawler]['picks'] += 1
                if is_winner:
                    stats['team_stats'][team]['brawlers'][brawler]['wins'] += 1
                
                stats['team_stats'][team]['players'][player_tag]['matches'] += 1
                stats['team_stats'][team]['players'][player_tag]['brawlers'][brawler]['picks'] += 1
                if is_winner:
                    stats['team_stats'][team]['players'][player_tag]['brawlers'][brawler]['wins'] += 1
                
                if match['star_player_tag'] == player_tag:
                    stats['team_stats'][team]['players'][player_tag]['star_player'] += 1
                
                stats['brawler_stats'][brawler]['picks'] += 1
                if is_winner:
                    stats['brawler_stats'][brawler]['wins'] += 1
                
                stats['map_stats'][mode][map_name]['teams'][team]['brawlers'][brawler]['picks'] += 1
                if is_winner:
                    stats['map_stats'][mode][map_name]['teams'][team]['brawlers'][brawler]['wins'] += 1
        
        stats['map_stats'][mode][map_name]['matches'] += 1
        stats['map_stats'][mode][map_name]['teams'][winner]['wins'] += 1
        stats['map_stats'][mode][map_name]['teams'][loser]['losses'] += 1
    
    return stats


def save_statistics():
    """Calculate and save enhanced statistics to Excel with region filtering"""
    if not matches_data:
        return
    
    stats = calculate_advanced_statistics()
    if not stats:
        return
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # ===== DASHBOARD SHEET WITH REGION FILTER =====
    ws_dashboard = wb.create_sheet("Dashboard", 0)
    
    # Title
    ws_dashboard['A1'] = 'Brawl Stars Team Statistics - Dashboard'
    ws_dashboard['A1'].font = Font(bold=True, size=16, color='FFFFFF')
    ws_dashboard['A1'].fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    ws_dashboard.merge_cells('A1:F1')
    ws_dashboard.row_dimensions[1].height = 25
    
    # Filter Instructions
    ws_dashboard['A3'] = '🔽 Select Region:'
    ws_dashboard['A3'].font = Font(bold=True, size=12)
    ws_dashboard['B3'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    
    # Region dropdown in B3
    region_options = ['ALL'] + CONFIG['REGIONS']
    dv_region = DataValidation(
        type="list",
        formula1=f'"{",".join(region_options)}"',
        allow_blank=False
    )
    ws_dashboard.add_data_validation(dv_region)
    dv_region.add('B3')
    ws_dashboard['B3'] = 'ALL'  # Default value
    
    # Instructions
    ws_dashboard['D3'] = 'NOTE: Dropdown filters data, but filtered-out teams leave blank rows (Excel limitation with formulas)'
    ws_dashboard['D3'].font = Font(italic=True, size=9)
    ws_dashboard.merge_cells('D3:F3')
    
    # Prepare ALL teams data sorted by wins
    all_teams = [(team, data) for team, data in stats['team_stats'].items()]
    all_teams.sort(key=lambda x: x[1]['wins'], reverse=True)
    
    # Group teams by region for better filtering
    teams_by_region = defaultdict(list)
    for team, data in all_teams:
        teams_by_region[data['region']].append((team, data))
    
    # Headers for display table
    display_headers = ['Team Name', 'Region', 'Matches', 'Wins', 'Losses', 'Win Rate']
    for col, header in enumerate(display_headers, 1):
        cell = ws_dashboard.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data starting at row 6 - STATIC approach with AutoFilter
    row = 6
    for team, data in all_teams:
        win_rate = (data['wins'] / data['matches']) if data['matches'] > 0 else 0
        
        ws_dashboard.cell(row=row, column=1, value=team)
        ws_dashboard.cell(row=row, column=2, value=data['region'])
        ws_dashboard.cell(row=row, column=3, value=data['matches'])
        ws_dashboard.cell(row=row, column=4, value=data['wins'])
        ws_dashboard.cell(row=row, column=5, value=data['losses'])
        ws_dashboard.cell(row=row, column=6, value=win_rate)
        ws_dashboard.cell(row=row, column=6).number_format = '0.0%'
        row += 1
    
    last_data_row = row - 1
    
    # Enable Excel's AutoFilter
    ws_dashboard.auto_filter.ref = f"A5:F{last_data_row}"
    
    # Set column widths for visible columns
    ws_dashboard.column_dimensions['A'].width = 25
    ws_dashboard.column_dimensions['B'].width = 12
    ws_dashboard.column_dimensions['C'].width = 12
    ws_dashboard.column_dimensions['D'].width = 12
    ws_dashboard.column_dimensions['E'].width = 12
    ws_dashboard.column_dimensions['F'].width = 14
    
    # Add summary statistics that count visible rows
    summary_start_row = last_data_row + 3
    ws_dashboard.cell(row=summary_start_row, column=1, value='Summary Statistics')
    ws_dashboard.cell(row=summary_start_row, column=1).font = Font(bold=True, size=12)
    
    ws_dashboard.cell(row=summary_start_row + 1, column=1, value='Total Teams:')
    ws_dashboard.cell(row=summary_start_row + 1, column=2, value=len(all_teams))
    ws_dashboard.cell(row=summary_start_row + 1, column=2).font = Font(bold=True)
    
    ws_dashboard.cell(row=summary_start_row + 2, column=1, value='Total Matches:')
    ws_dashboard.cell(row=summary_start_row + 2, column=2, 
                      value=sum(data['matches'] for _, data in all_teams))
    ws_dashboard.cell(row=summary_start_row + 2, column=2).font = Font(bold=True)
    
    # Add instruction for using AutoFilter
    ws_dashboard.cell(row=summary_start_row + 4, column=1, 
                      value='💡 TIP: Click the filter arrow in the "Region" column header to filter by region')
    ws_dashboard.cell(row=summary_start_row + 4, column=1).font = Font(italic=True, size=10, color='0066CC')
    ws_dashboard.merge_cells(f'A{summary_start_row + 4}:F{summary_start_row + 4}')
    
    # ===== REST OF THE SHEETS (keep existing code) =====
    
    # 1. Overall Stats by Region
    ws_overall = wb.create_sheet("Overall Stats")
    ws_overall['A1'] = 'Overall Team Statistics - All Regions'
    ws_overall['A1'].font = Font(bold=True, size=14)
    ws_overall.merge_cells('A1:F1')
    
    ws_overall['A2'] = 'Navigate to region-specific sheets (tabs at bottom) for detailed team breakdowns'
    ws_overall['A2'].font = Font(italic=True, size=10)
    ws_overall.merge_cells('A2:F2')
    
    headers = ['Team Name', 'Region', 'Matches', 'Wins', 'Losses', 'Win Rate']
    for col, header in enumerate(headers, 1):
        cell = ws_overall.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    
    row = 5
    for team, data in all_teams:
        win_rate = (data['wins'] / data['matches'] * 100) if data['matches'] > 0 else 0
        ws_overall.cell(row=row, column=1, value=team)
        ws_overall.cell(row=row, column=2, value=data['region'])
        ws_overall.cell(row=row, column=3, value=data['matches'])
        ws_overall.cell(row=row, column=4, value=data['wins'])
        ws_overall.cell(row=row, column=5, value=data['losses'])
        ws_overall.cell(row=row, column=6, value=f"{win_rate:.1f}%")
        row += 1
    
    ws_overall.column_dimensions['A'].width = 25
    ws_overall.column_dimensions['B'].width = 12
    ws_overall.column_dimensions['C'].width = 12
    ws_overall.column_dimensions['D'].width = 12
    ws_overall.column_dimensions['E'].width = 12
    ws_overall.column_dimensions['F'].width = 14
    
    # 2. Regional Stats (Detailed)
    for region in CONFIG['REGIONS']:
        ws_region = wb.create_sheet(f"{region} Teams")
        
        region_teams = [(team, data) for team, data in stats['team_stats'].items() if data['region'] == region]
        
        if not region_teams:
            ws_region['A1'] = f'No teams from {region} region'
            continue
        
        ws_region['A1'] = f'{region} Region - Team Statistics'
        ws_region['A1'].font = Font(bold=True, size=14)
        ws_region.merge_cells('A1:E1')
        
        ws_region['A2'] = f'Detailed statistics for all teams in the {region} region'
        ws_region['A2'].font = Font(italic=True, size=10)
        ws_region.merge_cells('A2:E2')
        
        row = 4
        for team, data in region_teams:
            ws_region.cell(row=row, column=1, value=f"Team: {team}").font = Font(bold=True, size=13)
            ws_region.cell(row=row, column=1).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            row += 1
            
            win_rate = (data['wins'] / data['matches'] * 100) if data['matches'] > 0 else 0
            ws_region.cell(row=row, column=1, value='Total Matches:')
            ws_region.cell(row=row, column=2, value=data['matches'])
            row += 1
            ws_region.cell(row=row, column=1, value='Wins:')
            ws_region.cell(row=row, column=2, value=data['wins'])
            row += 1
            ws_region.cell(row=row, column=1, value='Losses:')
            ws_region.cell(row=row, column=2, value=data['losses'])
            row += 1
            ws_region.cell(row=row, column=1, value='Win Rate:')
            ws_region.cell(row=row, column=2, value=f"{win_rate:.1f}%")
            ws_region.cell(row=row, column=2).font = Font(bold=True)
            row += 2
            
            ws_region.cell(row=row, column=1, value='Brawler Statistics:').font = Font(bold=True, size=11)
            row += 1
            ws_region.cell(row=row, column=1, value='Brawler')
            ws_region.cell(row=row, column=2, value='Picks')
            ws_region.cell(row=row, column=3, value='Wins')
            ws_region.cell(row=row, column=4, value='Win Rate')
            for col in range(1, 5):
                ws_region.cell(row=row, column=col).font = Font(bold=True)
            row += 1
            
            for brawler, bdata in sorted(data['brawlers'].items(), key=lambda x: x[1]['picks'], reverse=True):
                brawler_wr = (bdata['wins'] / bdata['picks'] * 100) if bdata['picks'] > 0 else 0
                ws_region.cell(row=row, column=1, value=brawler)
                ws_region.cell(row=row, column=2, value=bdata['picks'])
                ws_region.cell(row=row, column=3, value=bdata['wins'])
                ws_region.cell(row=row, column=4, value=f"{brawler_wr:.1f}%")
                row += 1
            
            row += 2
            
            ws_region.cell(row=row, column=1, value='Player Statistics:').font = Font(bold=True, size=11)
            row += 1
            
            for player_tag, pdata in data['players'].items():
                player_name = player_names_map.get(player_tag, player_tag)
                total_picks = sum(b['picks'] for b in pdata['brawlers'].values())
                total_wins = sum(b['wins'] for b in pdata['brawlers'].values())
                player_wr = (total_wins / total_picks * 100) if total_picks > 0 else 0
                star_rate = (pdata['star_player'] / pdata['matches'] * 100) if pdata['matches'] > 0 else 0
                
                ws_region.cell(row=row, column=1, value=f"Player: {player_name}").font = Font(bold=True)
                ws_region.cell(row=row, column=1).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                row += 1
                ws_region.cell(row=row, column=1, value='Matches:')
                ws_region.cell(row=row, column=2, value=pdata['matches'])
                row += 1
                ws_region.cell(row=row, column=1, value='Win Rate:')
                ws_region.cell(row=row, column=2, value=f"{player_wr:.1f}%")
                row += 1
                ws_region.cell(row=row, column=1, value='Star Player Rate:')
                ws_region.cell(row=row, column=2, value=f"{star_rate:.1f}%")
                row += 1
                
                ws_region.cell(row=row, column=1, value='Brawler')
                ws_region.cell(row=row, column=2, value='Picks')
                ws_region.cell(row=row, column=3, value='Wins')
                ws_region.cell(row=row, column=4, value='Win Rate')
                ws_region.cell(row=row, column=5, value='Pick Rate')
                for col in range(1, 6):
                    ws_region.cell(row=row, column=col).font = Font(italic=True)
                row += 1
                
                for brawler, bdata in sorted(pdata['brawlers'].items(), key=lambda x: x[1]['picks'], reverse=True):
                    brawler_wr = (bdata['wins'] / bdata['picks'] * 100) if bdata['picks'] > 0 else 0
                    pick_rate = (bdata['picks'] / total_picks * 100) if total_picks > 0 else 0
                    ws_region.cell(row=row, column=1, value=brawler)
                    ws_region.cell(row=row, column=2, value=bdata['picks'])
                    ws_region.cell(row=row, column=3, value=bdata['wins'])
                    ws_region.cell(row=row, column=4, value=f"{brawler_wr:.1f}%")
                    ws_region.cell(row=row, column=5, value=f"{pick_rate:.1f}%")
                    row += 1
                
                row += 2
            
            row += 3
        
        ws_region.column_dimensions['A'].width = 25
        ws_region.column_dimensions['B'].width = 12
        ws_region.column_dimensions['C'].width = 12
        ws_region.column_dimensions['D'].width = 12
        ws_region.column_dimensions['E'].width = 12
    
    # 3. Brawler Stats
    ws_brawlers = wb.create_sheet("Brawler Stats")
    ws_brawlers['A1'] = 'Global Brawler Statistics - All Teams'
    ws_brawlers['A1'].font = Font(bold=True, size=14)
    ws_brawlers.merge_cells('A1:E1')
    
    ws_brawlers['A2'] = 'Aggregate statistics across all matches and teams'
    ws_brawlers['A2'].font = Font(italic=True, size=10)
    ws_brawlers.merge_cells('A2:E2')
    
    ws_brawlers.cell(row=4, column=1, value='Brawler').font = Font(bold=True)
    ws_brawlers.cell(row=4, column=2, value='Picks').font = Font(bold=True)
    ws_brawlers.cell(row=4, column=3, value='Wins').font = Font(bold=True)
    ws_brawlers.cell(row=4, column=4, value='Win Rate').font = Font(bold=True)
    ws_brawlers.cell(row=4, column=5, value='Pick Rate').font = Font(bold=True)
    
    for col in range(1, 6):
        ws_brawlers.cell(row=4, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws_brawlers.cell(row=4, column=col).font = Font(color='FFFFFF', bold=True)
    
    total_picks = sum(b['picks'] for b in stats['brawler_stats'].values())
    row = 5
    for brawler, data in sorted(stats['brawler_stats'].items(), key=lambda x: x[1]['picks'], reverse=True):
        win_rate = (data['wins'] / data['picks'] * 100) if data['picks'] > 0 else 0
        pick_rate = (data['picks'] / total_picks * 100) if total_picks > 0 else 0
        ws_brawlers.cell(row=row, column=1, value=brawler)
        ws_brawlers.cell(row=row, column=2, value=data['picks'])
        ws_brawlers.cell(row=row, column=3, value=data['wins'])
        ws_brawlers.cell(row=row, column=4, value=f"{win_rate:.1f}%")
        ws_brawlers.cell(row=row, column=5, value=f"{pick_rate:.1f}%")
        row += 1
    
    ws_brawlers.column_dimensions['A'].width = 20
    ws_brawlers.column_dimensions['B'].width = 12
    ws_brawlers.column_dimensions['C'].width = 12
    ws_brawlers.column_dimensions['D'].width = 12
    ws_brawlers.column_dimensions['E'].width = 12
    
    # 4. Map Stats
    ws_map_summary = wb.create_sheet("Map Stats Summary")
    ws_map_summary['A1'] = 'Map Statistics Overview'
    ws_map_summary['A1'].font = Font(bold=True, size=14)
    ws_map_summary.merge_cells('A1:D1')
    
    ws_map_summary['A2'] = 'Navigate to individual mode sheets for detailed map breakdowns'
    ws_map_summary['A2'].font = Font(italic=True, size=10)
    ws_map_summary.merge_cells('A2:D2')
    
    summary_row = 4
    ws_map_summary.cell(row=summary_row, column=1, value='Mode').font = Font(bold=True)
    ws_map_summary.cell(row=summary_row, column=2, value='Total Maps').font = Font(bold=True)
    ws_map_summary.cell(row=summary_row, column=3, value='Total Matches').font = Font(bold=True)
    ws_map_summary.cell(row=summary_row, column=4, value='Sheet Name').font = Font(bold=True)
    
    for col in range(1, 5):
        ws_map_summary.cell(row=summary_row, column=col).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws_map_summary.cell(row=summary_row, column=col).font = Font(color='FFFFFF', bold=True)
    
    summary_row += 1
    
    for mode in CONFIG['MODES']:
        if mode not in stats['map_stats']:
            continue
        
        total_maps = len(stats['map_stats'][mode])
        total_matches = sum(map_data['matches'] for map_data in stats['map_stats'][mode].values())
        ws_map_summary.cell(row=summary_row, column=1, value=mode)
        ws_map_summary.cell(row=summary_row, column=2, value=total_maps)
        ws_map_summary.cell(row=summary_row, column=3, value=total_matches)
        ws_map_summary.cell(row=summary_row, column=4, value=f"See '{mode[:31]}' tab")
        summary_row += 1
        
        sheet_name = mode[:31]
        ws_mode = wb.create_sheet(sheet_name)
        
        ws_mode['A1'] = f'{mode} - Map Statistics'
        ws_mode['A1'].font = Font(bold=True, size=14)
        ws_mode.merge_cells('A1:F1')
        
        ws_mode['A2'] = f'All maps played in {mode} mode with team performance breakdowns'
        ws_mode['A2'].font = Font(italic=True, size=10)
        ws_mode.merge_cells('A2:F2')
        
        current_row = 4
        
        for map_name, map_data in sorted(stats['map_stats'][mode].items(), key=lambda x: x[1]['matches'], reverse=True):
            ws_mode.cell(row=current_row, column=1, value=f"Map: {map_name}").font = Font(bold=True, size=13)
            ws_mode.cell(row=current_row, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            ws_mode.cell(row=current_row, column=1).font = Font(bold=True, size=13, color='FFFFFF')
            ws_mode.merge_cells(f'A{current_row}:F{current_row}')
            current_row += 1
            
            ws_mode.cell(row=current_row, column=1, value='Total Matches:')
            ws_mode.cell(row=current_row, column=2, value=map_data['matches'])
            ws_mode.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 2
            
            ws_mode.cell(row=current_row, column=1, value='Overall Statistics').font = Font(bold=True, size=11)
            current_row += 1
            
            headers = ['Team', 'Wins', 'Losses', 'Win Rate', 'Total Matches']
            for col, header in enumerate(headers, 1):
                cell = ws_mode.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
            current_row += 1
            
            for team, team_data in sorted(map_data['teams'].items(), key=lambda x: x[1]['wins'], reverse=True):
                total = team_data['wins'] + team_data['losses']
                win_rate = (team_data['wins'] / total * 100) if total > 0 else 0
                ws_mode.cell(row=current_row, column=1, value=team)
                ws_mode.cell(row=current_row, column=2, value=team_data['wins'])
                ws_mode.cell(row=current_row, column=3, value=team_data['losses'])
                ws_mode.cell(row=current_row, column=4, value=f"{win_rate:.1f}%")
                ws_mode.cell(row=current_row, column=5, value=total)
                current_row += 1
            
            current_row += 2
            
            ws_mode.cell(row=current_row, column=1, value='Detailed Team Performance').font = Font(bold=True, size=11)
            current_row += 1
            
            for team, team_data in sorted(map_data['teams'].items(), key=lambda x: x[1]['wins'], reverse=True):
                total = team_data['wins'] + team_data['losses']
                win_rate = (team_data['wins'] / total * 100) if total > 0 else 0
                
                ws_mode.cell(row=current_row, column=1, value=f"━━━ {team} ━━━").font = Font(bold=True, size=11)
                ws_mode.cell(row=current_row, column=1).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
                ws_mode.merge_cells(f'A{current_row}:F{current_row}')
                current_row += 1
                
                ws_mode.cell(row=current_row, column=1, value='Record:')
                ws_mode.cell(row=current_row, column=2, value=f"{team_data['wins']}-{team_data['losses']}")
                ws_mode.cell(row=current_row, column=3, value='Win Rate:')
                ws_mode.cell(row=current_row, column=4, value=f"{win_rate:.1f}%")
                ws_mode.cell(row=current_row, column=4).font = Font(bold=True)
                current_row += 2
                
                if team_data['brawlers']:
                    ws_mode.cell(row=current_row, column=1, value='Brawler Statistics:').font = Font(bold=True)
                    current_row += 1
                    
                    headers = ['Brawler', 'Picks', 'Wins', 'Win Rate', 'Pick %']
                    for col, header in enumerate(headers, 1):
                        cell = ws_mode.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                    current_row += 1
                    
                    total_brawler_picks = sum(b['picks'] for b in team_data['brawlers'].values())
                    for brawler, bdata in sorted(team_data['brawlers'].items(), key=lambda x: x[1]['picks'], reverse=True):
                        b_wr = (bdata['wins'] / bdata['picks'] * 100) if bdata['picks'] > 0 else 0
                        b_pr = (bdata['picks'] / total_brawler_picks * 100) if total_brawler_picks > 0 else 0
                        ws_mode.cell(row=current_row, column=1, value=brawler)
                        ws_mode.cell(row=current_row, column=2, value=bdata['picks'])
                        ws_mode.cell(row=current_row, column=3, value=bdata['wins'])
                        ws_mode.cell(row=current_row, column=4, value=f"{b_wr:.1f}%")
                        ws_mode.cell(row=current_row, column=5, value=f"{b_pr:.1f}%")
                        current_row += 1
                
                current_row += 2
            
            current_row += 3
        
        ws_mode.column_dimensions['A'].width = 25
        ws_mode.column_dimensions['B'].width = 12
        ws_mode.column_dimensions['C'].width = 12
        ws_mode.column_dimensions['D'].width = 14
        ws_mode.column_dimensions['E'].width = 12
        ws_mode.column_dimensions['F'].width = 12
    
    ws_map_summary.column_dimensions['A'].width = 20
    ws_map_summary.column_dimensions['B'].width = 15
    ws_map_summary.column_dimensions['C'].width = 18
    ws_map_summary.column_dimensions['D'].width = 30
    
    wb.save(CONFIG['STATS_FILE'])
    print(f"Saved enhanced statistics to {CONFIG['STATS_FILE']}")
    print(f"  - Dashboard sheet with region filter created")
    print(f"  - Created {len([s for s in wb.sheetnames if s in CONFIG['MODES']])} mode-specific sheets")


def run_analysis():
    """Wrapper function for scheduled analysis"""
    try:
        analyze_matches()
    except Exception as e:
        print(f"Error during analysis: {e}")


def main():
    """Main function"""
    print("Brawl Stars Match Analyzer - 24/7 Service")
    print("=" * 60)
    
    load_existing_matches()
    run_analysis()
    
    schedule.every(CONFIG['CHECK_INTERVAL_MINUTES']).minutes.do(run_analysis)
    
    print(f"\nScheduled to check every {CONFIG['CHECK_INTERVAL_MINUTES']} minutes")
    print("Service running. Press Ctrl+C to stop.\n")
    
    while True:
        schedule.run_pending()
        time.sleep(60)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nService stopped by user")
    except Exception as e:
        print(f"\nFatal error: {e}")